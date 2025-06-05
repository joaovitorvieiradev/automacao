import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side

# Muda o diretório para a pasta onde está o arquivo
pasta_codigo = os.getcwd()
pasta_destino = os.path.join(pasta_codigo, "tratamento_de_planilhas")
os.chdir(pasta_destino)

# Lê a planilha
df = pd.read_excel("BaseDeDados.xlsx")

# Limpeza: remove linhas inválidas
df = df[df['Curso'].notna()]
df = df[~df['Curso'].astype(str).str.contains("resultados encontrados", case=False)]

# Cria a tabela dinâmica
tabela_dinamica = pd.pivot_table(
    df,
    index='Curso',
    columns='Turno',
    aggfunc='size',
    fill_value=0
)

# Garante que as colunas Matutino e Vespertino existam (mesmo que não estejam na base original)
for turno in ['Matutino', 'Vespertino']:
    if turno not in tabela_dinamica.columns:
        tabela_dinamica[turno] = 0

# Remove outras colunas indesejadas (colunas que não sejam Matutino ou Vespertino)
tabela_dinamica = tabela_dinamica[['Matutino', 'Vespertino']]

# Adiciona coluna Total
tabela_dinamica['Total'] = tabela_dinamica[['Matutino', 'Vespertino']].sum(axis=1)

# Define o caminho absoluto para o TXT de filtros
caminho_filtros = os.path.join(pasta_codigo, "automacao", "filtros_selecionados.txt")

with open(caminho_filtros, 'r', encoding='utf-8') as f:
    linhas = f.readlines()

trimestre_raw = linhas[1].strip()  # linha 2 do arquivo

if trimestre_raw.lower() == "todos":
    trimestre = "todos"
else:
    trimestre = trimestre_raw.split('/')[0]  # extrai "1" de "1/23", por exemplo

# Define para qual célula colar os dados
trimestre_map = {
    '1': 'B8',
    '2': 'H8',
    '3': 'B40',
    '4': 'H40',
    'todos': 'B65'
}

celula_destino = trimestre_map.get(trimestre)
if not celula_destino:
    raise ValueError(f"Trimestre inválido: {trimestre_raw}")

# === Parte 3: Abrir arquivo .xlsm e colar a tabela na célula certa ===

# Caminho completo para a planilha de destino
caminho_arquivo = os.path.join(pasta_codigo, "tratamento_de_planilhas", "MAPA DE ALUNOS.xlsm")
wb = load_workbook(caminho_arquivo, keep_vba=True)
ws = wb.active  # ou especifique uma aba: wb['NomeDaAba']

# Localiza célula de início (exemplo: B8 = coluna 2, linha 8)
col_ini = ws[celula_destino].column
row_ini = ws[celula_destino].row

# Cola a tabela dinâmica na planilha
for i, (index, row) in enumerate(tabela_dinamica.iterrows()):
    ws.cell(row=row_ini + i, column=col_ini, value=index)  # Coluna do nome do curso
    for j, value in enumerate(row):
        ws.cell(row=row_ini + i, column=col_ini + j + 1, value=value)  # Colunas Matutino, Vespertino, Total

# Define borda fina preta
borda = Border(
    left=Side(border_style="thin", color="000000"),
    right=Side(border_style="thin", color="000000"),
    top=Side(border_style="thin", color="000000"),
    bottom=Side(border_style="thin", color="000000")
)

num_linhas = tabela_dinamica.shape[0]
num_colunas = tabela_dinamica.shape[1] + 1  # +1 para a coluna do índice

# Aplica bordas em todas as células da tabela colada
for i in range(num_linhas):
    for j in range(num_colunas):
        ws.cell(row=row_ini + i, column=col_ini + j).border = borda

# Salva e abre o arquivo
wb.save(caminho_arquivo)
os.startfile(caminho_arquivo)
